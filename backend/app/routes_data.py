from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from app.security import verify_token
from app.database import get_supabase
from typing import Optional
from datetime import datetime, timedelta
import io
import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api/data", tags=["data"])
security = HTTPBearer()

BATCH_SIZE = 1000


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token_data = verify_token(credentials.credentials)
        return token_data.username
    except:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")


def fetch_all_via_rpc(supabase, rpc_name: str, params: dict) -> list:
    all_data = []
    offset = 0
    while True:
        batch_params = {**params, "p_limit": BATCH_SIZE, "p_offset": offset}
        result = supabase.rpc(rpc_name, batch_params).execute()
        batch = result.data or []
        all_data.extend(batch)
        if len(batch) < BATCH_SIZE:
            break
        offset += BATCH_SIZE
    return all_data


@router.get("/harga-beras")
async def get_harga_beras(
    username: str = Depends(get_current_user),
    tipe_harga_id: Optional[int] = Query(None),
    variant_id: Optional[int] = Query(None),
    tanggal_start: Optional[str] = Query(None),
    tanggal_end: Optional[str] = Query(None),
):
    """
    Ambil data harga beras dengan filter tipe harga, variant, dan rentang tanggal.
    Mendukung pagination otomatis per 1000 baris via RPC.
    """
    try:
        supabase = get_supabase()
        params = {}
        if tipe_harga_id is not None:
            params["p_tipe_harga_id"] = tipe_harga_id
        if variant_id is not None:
            params["p_variant_id"] = variant_id
        if tanggal_start:
            params["p_tanggal_start"] = tanggal_start
        if tanggal_end:
            params["p_tanggal_end"] = tanggal_end

        data = fetch_all_via_rpc(supabase, "get_harga_beras_table", params)
        return {"status": "success", "data": data, "total": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/harga-beras/export")
async def export_harga_beras(
    username: str = Depends(get_current_user),
    tipe_harga_id: Optional[int] = Query(None),
    variant_id: Optional[int] = Query(None),
    tanggal_start: Optional[str] = Query(None),
    tanggal_end: Optional[str] = Query(None),
):
    """Export data harga beras ke CSV."""
    try:
        supabase = get_supabase()
        params = {}
        if tipe_harga_id is not None:
            params["p_tipe_harga_id"] = tipe_harga_id
        if variant_id is not None:
            params["p_variant_id"] = variant_id
        if tanggal_start:
            params["p_tanggal_start"] = tanggal_start
        if tanggal_end:
            params["p_tanggal_end"] = tanggal_end

        data = fetch_all_via_rpc(supabase, "get_harga_beras_table", params)

        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=harga_beras.csv"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/harga-beras/export-excel")
async def export_harga_beras_excel(
    username: str = Depends(get_current_user),
    tipe_harga_id: Optional[int] = Query(None),
    variant_id: Optional[int] = Query(None),
    tanggal_start: Optional[str] = Query(None),
    tanggal_end: Optional[str] = Query(None),
    tipe_harga_nama: Optional[str] = Query(None),
    variant_nama: Optional[str] = Query(None),
):
    """Export data harga beras ke Excel dengan 2 sheet: analisa dan data mentah."""
    try:
        supabase = get_supabase()
        params = {}
        if tipe_harga_id is not None:
            params["p_tipe_harga_id"] = tipe_harga_id
        if variant_id is not None:
            params["p_variant_id"] = variant_id
        if tanggal_start:
            params["p_tanggal_start"] = tanggal_start
        if tanggal_end:
            params["p_tanggal_end"] = tanggal_end

        # Ambil data filter (range tanggal terpilih)
        data = fetch_all_via_rpc(supabase, "get_harga_beras_table", params)

        # Ambil data bulan lalu (untuk sheet data mentah)
        if tanggal_end:
            end_dt = datetime.strptime(tanggal_end, "%Y-%m-%d")
        else:
            end_dt = datetime.today()
        bulan_lalu_start = (end_dt.replace(day=1) - timedelta(days=1)).replace(day=1)
        bulan_lalu_end   = end_dt.replace(day=1) - timedelta(days=1)

        params_bulan_lalu = {
            "p_tanggal_start": bulan_lalu_start.strftime("%Y-%m-%d"),
            "p_tanggal_end":   bulan_lalu_end.strftime("%Y-%m-%d"),
        }
        if tipe_harga_id is not None:
            params_bulan_lalu["p_tipe_harga_id"] = tipe_harga_id
        if variant_id is not None:
            params_bulan_lalu["p_variant_id"] = variant_id
        data_bulan_lalu = fetch_all_via_rpc(supabase, "get_harga_beras_table", params_bulan_lalu)

        # ── Helper style ─────────────────────────────────────────────────────
        BLUE_DARK  = PatternFill("solid", fgColor="1E40AF")  # blue-800
        BLUE_MID   = PatternFill("solid", fgColor="1D4ED8")  # blue-700
        BLUE_LIGHT = PatternFill("solid", fgColor="EFF6FF")  # blue-50
        WHITE      = PatternFill("solid", fgColor="FFFFFF")
        GRAY_50    = PatternFill("solid", fgColor="F9FAFB")
        WHITE_FONT = Font(color="FFFFFF", bold=True, size=10)
        BOLD       = Font(bold=True, size=10)
        NORMAL     = Font(size=10)
        BLUE_FONT  = Font(bold=True, color="1E3A5F", size=10)
        thin       = Side(style="thin", color="CBD5E1")
        BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
        CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT       = Alignment(horizontal="left",   vertical="center")

        def style(cell, fill=None, font=None, align=CENTER, border=BORDER):
            if fill:   cell.fill   = fill
            if font:   cell.font   = font
            if align:  cell.alignment = align
            if border: cell.border = border

        wb = openpyxl.Workbook()

        # ════════════════════════════════════════════════════════════════════
        # SHEET 1: analisa (sama persis dengan tampilan HTML)
        # ════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "analisa"

        # Pivot data filter: kode → {tanggal: harga, bulan_lalu, nama}
        pivot = {}
        tanggal_set = set()
        for row in data:
            k = row["kode_kab_kota"]
            t = row["tanggal"]
            tanggal_set.add(t)
            if k not in pivot:
                pivot[k] = {"nama": row.get("nama_kab_kota", k), "harga": {}, "bulan_lalu": row.get("harga_rata_bulan_lalu")}
            pivot[k]["harga"][t] = row.get("harga_rata_tanggal")
            pivot[k]["bulan_lalu"] = row.get("harga_rata_bulan_lalu")

        tanggal_list = sorted(tanggal_set)

        # Urutkan kab berdasarkan |pct| desc, Jawa Timur ke bawah
        def get_pct(entry):
            bl = entry["bulan_lalu"]
            vals = [v for v in entry["harga"].values() if v is not None]
            if not vals or not bl:
                return 0
            rata = sum(vals) / len(vals)
            return abs((rata - bl) / bl * 100)

        kab_sorted = sorted(pivot.items(), key=lambda x: get_pct(x[1]), reverse=True)
        kab_jatim  = [(k, v) for k, v in kab_sorted if "jawa timur" in v["nama"].lower()]
        kab_biasa  = [(k, v) for k, v in kab_sorted if "jawa timur" not in v["nama"].lower()]

        n_tgl = len(tanggal_list)
        total_cols = 2 + n_tgl + 2  # Kab/Kota | Rata BulanLalu | tgl... | Rata Terpilih | %

        # Header row 1
        ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws1.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        ws1.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + n_tgl)
        ws1.merge_cells(start_row=1, start_column=3 + n_tgl, end_row=2, end_column=3 + n_tgl)
        ws1.merge_cells(start_row=1, start_column=4 + n_tgl, end_row=2, end_column=4 + n_tgl)

        h1 = ws1.cell(1, 1, "Kab/Kota");            style(h1, BLUE_DARK, WHITE_FONT)
        h2 = ws1.cell(1, 2, "Rata-rata Harga Bulan Lalu"); style(h2, BLUE_DARK, WHITE_FONT)
        h3 = ws1.cell(1, 3, "Tanggal");              style(h3, BLUE_DARK, WHITE_FONT)
        h4 = ws1.cell(1, 3 + n_tgl, "Rata-rata Harga Terpilih"); style(h4, BLUE_DARK, WHITE_FONT)
        h5 = ws1.cell(1, 4 + n_tgl, "vs Rata-rata Bulan Lalu (%)"); style(h5, BLUE_DARK, WHITE_FONT)

        # Header row 2 — tanggal
        for i, tgl in enumerate(tanggal_list):
            dt = datetime.strptime(tgl, "%Y-%m-%d")
            label = dt.strftime("%d %b '%y")
            c = ws1.cell(2, 3 + i, label)
            style(c, BLUE_MID, WHITE_FONT)

        # Data rows
        for ridx, (kode, entry) in enumerate(kab_biasa):
            r = 3 + ridx
            fill = WHITE if ridx % 2 == 0 else GRAY_50
            vals = [entry["harga"].get(t) for t in tanggal_list]
            valid = [v for v in vals if v is not None]
            rata  = sum(valid) / len(valid) if valid else None
            bl    = entry["bulan_lalu"]
            pct   = (rata - bl) / bl * 100 if rata and bl else None

            c = ws1.cell(r, 1, entry["nama"]); style(c, fill, NORMAL, LEFT)
            c = ws1.cell(r, 2, bl);            style(c, fill, NORMAL); ws1.cell(r, 2).number_format = '#,##0'
            for i, v in enumerate(vals):
                c = ws1.cell(r, 3 + i, v); style(c, fill, NORMAL); c.number_format = '#,##0'
            c = ws1.cell(r, 3 + n_tgl, rata); style(c, fill, BOLD); c.number_format = '#,##0'
            if pct is not None:
                c = ws1.cell(r, 4 + n_tgl, round(pct, 2))
                c.number_format = '+0.00%;-0.00%'
                c.font = Font(bold=True, color="DC2626" if pct > 0 else "16A34A", size=10)
            else:
                c = ws1.cell(r, 4 + n_tgl, "-")
            style(c, fill, border=BORDER); c.alignment = CENTER

        # Footer row — Jawa Timur
        for kode, entry in kab_jatim:
            r = 3 + len(kab_biasa)
            vals = [entry["harga"].get(t) for t in tanggal_list]
            valid = [v for v in vals if v is not None]
            rata  = sum(valid) / len(valid) if valid else None
            bl    = entry["bulan_lalu"]
            pct   = (rata - bl) / bl * 100 if rata and bl else None

            c = ws1.cell(r, 1, entry["nama"]); style(c, BLUE_LIGHT, BLUE_FONT, LEFT)
            c = ws1.cell(r, 2, bl);            style(c, BLUE_LIGHT, BLUE_FONT); c.number_format = '#,##0'
            for i, v in enumerate(vals):
                c = ws1.cell(r, 3 + i, v); style(c, BLUE_LIGHT, BLUE_FONT); c.number_format = '#,##0'
            c = ws1.cell(r, 3 + n_tgl, rata); style(c, BLUE_LIGHT, BLUE_FONT); c.number_format = '#,##0'
            if pct is not None:
                c = ws1.cell(r, 4 + n_tgl, round(pct, 2))
                c.number_format = '+0.00%;-0.00%'
                c.font = Font(bold=True, color="DC2626" if pct > 0 else "16A34A", size=10)
            else:
                c = ws1.cell(r, 4 + n_tgl, "-")
            style(c, BLUE_LIGHT, border=BORDER); c.alignment = CENTER

        # Lebar kolom sheet analisa
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 24
        for i in range(n_tgl):
            ws1.column_dimensions[get_column_letter(3 + i)].width = 14
        ws1.column_dimensions[get_column_letter(3 + n_tgl)].width = 22
        ws1.column_dimensions[get_column_letter(4 + n_tgl)].width = 22
        ws1.row_dimensions[1].height = 30
        ws1.row_dimensions[2].height = 22

        # ════════════════════════════════════════════════════════════════════
        # SHEET 2: data mentah — pivot per tanggal, kolom = kab/kota
        # ════════════════════════════════════════════════════════════════════
        vname = (variant_nama or "variant").replace(" ", "_")
        tname = (tipe_harga_nama or "tipe").replace(" ", "_")
        ws2 = wb.create_sheet(title=f"data {vname} - {tname}"[:31])

        # Gabung data bulan lalu + data filter, urutkan tanggal
        all_data = data_bulan_lalu + data
        pivot2 = {}   # tanggal → {kode: harga}
        kab_map = {}  # kode → nama
        for row in all_data:
            k = row["kode_kab_kota"]
            t = row["tanggal"]
            kab_map[k] = row.get("nama_kab_kota", k)
            if t not in pivot2:
                pivot2[t] = {}
            pivot2[t][k] = row.get("harga_rata_tanggal")

        all_tanggal = sorted(pivot2.keys())

        # Urutan kolom: Jawa Timur di akhir
        kab_order_all = [k for k in kab_map if "jawa timur" not in kab_map[k].lower()]
        kab_order_jt  = [k for k in kab_map if "jawa timur" in kab_map[k].lower()]

        # Header
        ws2.cell(1, 1, "Tanggal");   style(ws2.cell(1,1), BLUE_DARK, WHITE_FONT, LEFT)
        col = 2
        for k in kab_order_all + kab_order_jt:
            c = ws2.cell(1, col, kab_map[k]); style(c, BLUE_DARK, WHITE_FONT)
            col += 1
        ws2.cell(1, col, "Minggu ke"); style(ws2.cell(1, col), BLUE_DARK, WHITE_FONT)

        # Data rows
        for ridx, tgl in enumerate(all_tanggal):
            r = 2 + ridx
            fill = WHITE if ridx % 2 == 0 else GRAY_50
            dt = datetime.strptime(tgl, "%Y-%m-%d")
            # Minggu ke dalam bulan
            minggu_ke = (dt.day - 1) // 7 + 1

            c = ws2.cell(r, 1, dt.strftime("%d/%m/%Y")); style(c, fill, NORMAL, LEFT)
            col = 2
            for k in kab_order_all + kab_order_jt:
                v = pivot2[tgl].get(k)
                c = ws2.cell(r, col, v); style(c, fill, NORMAL)
                if v is not None:
                    c.number_format = '#,##0'
                col += 1
            c = ws2.cell(r, col, f"Minggu {minggu_ke}"); style(c, fill, NORMAL)

        # Lebar kolom sheet data
        ws2.column_dimensions["A"].width = 14
        for i in range(len(kab_order_all + kab_order_jt) + 1):
            ws2.column_dimensions[get_column_letter(2 + i)].width = 16
        ws2.row_dimensions[1].height = 28

        # Output
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        fname = f"harga_beras_{vname}_{tname}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/variant")
async def get_variant(username: str = Depends(get_current_user)):
    """Daftar variant dari tabel variant via RPC."""
    try:
        supabase = get_supabase()
        result = supabase.rpc("get_variant", {}).execute()
        return {"status": "success", "data": result.data or []}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/list", dependencies=[Depends(get_current_user)])
async def list_data():
    return {"status": "success", "data": []}


@router.post("/upload", dependencies=[Depends(get_current_user)])
async def upload_data(file: UploadFile = File(...)):
    return {"status": "success", "message": f"File {file.filename} uploaded successfully"}


@router.delete("/delete/{data_id}", dependencies=[Depends(get_current_user)])
async def delete_data(data_id: str):
    return {"status": "success", "message": f"Data {data_id} deleted successfully"}
